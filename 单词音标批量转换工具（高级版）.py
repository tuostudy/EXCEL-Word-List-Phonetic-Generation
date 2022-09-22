import requests
import os
from lxml import etree
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Side, Border
from openpyxl import load_workbook


def get_phonetic():
    print('————' * 10)
    print('欢迎来到单词音标批量自动生成小程序2.0！')
    print('注意事项：\n'
          '1.请确保你的网络通畅！\n'
          '2.请确保你的文件后缀格式为xlsx而非csv!\n'
          '3.请确保你的单词全部在第一列!\n'
          '4.运行过程漫长请耐心等待，不要中途退出，否则不会得到任何结果!\n'
          '5.有什么不懂的地方，欢迎私信联系我，B站搜：TUO图欧君\n'
          )

    while True:  # 用户设置阶段
        try:
            file = input('请输入你的excel文件所在目录和文件名（比如：F:/我的单词本）：') + '.xlsx'  # 输入你的excel表格的文件存储位置
            if not os.path.exists(file):  # 判断文件是否存在
                print('你输入的文件路径有误，请重新输入！')
            else:
                print('————' * 10)
                print('————文件导入成功！————')
                break
        except NameError:
            print('你输入的文件路径有误，请重新输入！')

    workbook = load_workbook(file)  # 导入excel表格
    worksheet = workbook['Sheet1']  # 读取excel中的sheet1这张表

    worksheet.insert_cols(idx=2, amount=2)  # 向右插入两列准备存放数据，idx:插入列的位置，amount:插入的列数
    worksheet.insert_rows(idx=0, amount=1)  # 在顶部插入一行写入注释，idx:插入行的位置，amount:插入的行数
    worksheet['A1'] = '单词'  # 如果你的单词本第一行原本就有这些注释就不用这代码了，直接删掉就行
    worksheet['B1'] = '英音'
    worksheet['C1'] = '美音'
    worksheet['D1'] = '释义'

    font = Font(name='微软雅黑', size=20, bold=False)  # 设置全局字体大小样式
    worksheet['A1'].font = font
    worksheet['B1'].font = font
    worksheet['C1'].font = font
    worksheet['D1'].font = font

    alignment = Alignment(horizontal="left")  # 设置单元格对齐方式，用于修复墨墨词库csv文件顶部两处单元格的对齐异常问题
    worksheet['A2'].alignment = alignment  # 如果不需要此代码可删除
    worksheet['D2'].alignment = alignment  # 如果不需要此代码可删除
    border = Border(Side(style=None))  # 设置单元格边框样式，用于修复墨墨词库csv文件顶部两处单元格的边框异常问题
    worksheet['A2'].border = border  # 如果不需要此代码可删除
    worksheet['D2'].border = border  # 如果不需要此代码可删除

    worksheet.column_dimensions['A'].width = 30  # 设置全局列宽
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 30
    workbook.save(file)

    row_index = 2  # 默认结果放到第二行开始

    for row in worksheet.iter_rows(min_row=2, max_col=3):  # 以行迭代(最小第二行,最多第3列)
        word = row[0].value  # 获取第一列所有表格数据
        # 请确保你的单词在第1列，如果你的单词在第N列，请将代码换成：word = row[0+N].value，还要记得改其他各处代码中的数字！
        url = 'https://www.youdao.com/w/eng/{}'.format(word)  # 从有道获取音标
        try:
            data = requests.get(url).text
            html = etree.HTML(data)
            num1 = 'A' + str(row_index)
            num2 = 'B' + str(row_index)
            num3 = 'C' + str(row_index)
            num4 = 'D' + str(row_index)

            British_pron = html.xpath('//*[@id="phrsListTab"]/h2/div/span[1]/span/text()')[0]
            American_pron = html.xpath('//*[@id="phrsListTab"]/h2/div/span[2]/span/text()')[0]
            print('正在输出：' + British_pron, American_pron)

            worksheet.cell(row=row_index, column=2).value = British_pron  # 默认将英式音标的结果放到第2列
            worksheet.cell(row=row_index, column=3).value = American_pron  # 默认将美式音标的结果放到第3列

            worksheet[num1].font = font
            worksheet[num2].font = font
            worksheet[num3].font = font
            worksheet[num4].font = font
        except Exception as e:
            print(e, word)
            num1 = 'A' + str(row_index)
            num4 = 'D' + str(row_index)
            worksheet[num1].font = font
            worksheet[num4].font = font

        row_index += 1
    workbook.save(file)
    print("单词音标已经全部转换完毕！已经成功保存在原文件：" + file)


if __name__ == '__main__':
    get_phonetic()
