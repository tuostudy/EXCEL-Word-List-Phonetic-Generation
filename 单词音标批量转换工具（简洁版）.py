import requests
from lxml import etree
from openpyxl import load_workbook


def gainPron():
    print('欢迎来到单词音标批量自动生成小程序1.0！')
    file = input('请输入你的excel文件路径（比如：F:/我的单词本.xlsx）')  # 输入你的excel表格的文件存储位置
    workbook = load_workbook(file)  # 导入excel表格
    worksheet = workbook['Sheet1']  # 读取excel中的sheet1这张表
    row_index = 2   # 默认从第二行开始，如果你先想从第一行开始，请改为1
    for row in worksheet.iter_rows(min_row=2, max_col=3):  # 以行迭代(最小第二行,最多第3列)
        word = row[0].value
        # 请确保你的单词在第1列，如果你的单词在第N列，请将代码换成：word = row[0+N].value，还要记得改其他各处代码中的数字！
        url = 'https://www.youdao.com/w/eng/{}'.format(word)  # 从有道获取音标
        try:
            data = requests.get(url).text
            html = etree.HTML(data)

            British_pron = html.xpath('//*[@id="phrsListTab"]/h2/div/span[1]/span/text()')[0]
            American_pron = html.xpath('//*[@id="phrsListTab"]/h2/div/span[2]/span/text()')[0]
            print(British_pron, American_pron)

            worksheet.cell(row=row_index, column=2).value = British_pron  # 默认将英式音标的结果放到第2列
            worksheet.cell(row=row_index, column=3).value = American_pron  # 默认将美式音标的结果放到第3列
        except Exception as e:
            print(e, word)

        row_index += 1
    workbook.save(file)


if __name__ == '__main__':
    gainPron()
    print("单词音标已经全部转换完毕！")
